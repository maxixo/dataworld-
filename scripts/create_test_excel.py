#!/usr/bin/env python3
"""
Script to create a custom test Excel document with sample data
"""
import openpyxl
from datetime import datetime, timedelta
import random

def create_test_excel():
    """Create a test Excel file with sample data"""
    
    # Create a new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Data"
    
    # Add headers
    headers = [
        "Order ID",
        "Customer Name",
        "Product Category",
        "Product Name",
        "Quantity",
        "Unit Price",
        "Total Amount",
        "Order Date",
        "Region",
        "Status",
        "Payment Method",
        "Shipping Cost"
    ]
    
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)
        # Style the header row
        ws.cell(row=1, column=col_num).font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        ws.cell(row=1, column=col_num).fill = openpyxl.styles.PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
    
    # Sample data
    categories = ["Electronics", "Clothing", "Home & Garden", "Sports", "Books"]
    products = {
        "Electronics": ["Laptop", "Smartphone", "Tablet", "Headphones", "Camera"],
        "Clothing": ["T-Shirt", "Jeans", "Jacket", "Sneakers", "Watch"],
        "Home & Garden": ["Blender", "Coffee Maker", "Lamp", "Plant Pot", "Tool Set"],
        "Sports": ["Tennis Racket", "Yoga Mat", "Basketball", "Running Shoes", "Gym Bag"],
        "Books": ["Novel", "Textbook", "Magazine", "Comics", "Audiobook"]
    }
    regions = ["North", "South", "East", "West", "Central"]
    statuses = ["Completed", "Pending", "Shipped", "Delivered", "Cancelled"]
    payment_methods = ["Credit Card", "PayPal", "Bank Transfer", "Cash on Delivery"]
    customer_names = [
        "John Smith", "Emily Johnson", "Michael Brown", "Sarah Davis", "David Wilson",
        "Jessica Martinez", "Christopher Lee", "Amanda Taylor", "Daniel Anderson", "Ashley Thomas",
        "Matthew Jackson", "Jennifer White", "Andrew Harris", "Michelle Martin", "Joshua Thompson",
        "Laura Garcia", "James Robinson", "Stephanie Clark", "Robert Rodriguez", "Nicole Lewis"
    ]
    
    # Generate 50 sample records
    base_date = datetime(2024, 1, 1)
    
    for row_num in range(2, 52):  # 50 data rows
        order_id = f"ORD-{row_num:05d}"
        customer_name = random.choice(customer_names)
        category = random.choice(categories)
        product_name = random.choice(products[category])
        quantity = random.randint(1, 10)
        unit_price = round(random.uniform(10.00, 500.00), 2)
        total_amount = round(quantity * unit_price, 2)
        order_date = base_date + timedelta(days=random.randint(0, 364))
        region = random.choice(regions)
        status = random.choice(statuses)
        payment_method = random.choice(payment_methods)
        shipping_cost = round(random.uniform(5.00, 25.00), 2)
        
        ws.cell(row=row_num, column=1, value=order_id)
        ws.cell(row=row_num, column=2, value=customer_name)
        ws.cell(row=row_num, column=3, value=category)
        ws.cell(row=row_num, column=4, value=product_name)
        ws.cell(row=row_num, column=5, value=quantity)
        ws.cell(row=row_num, column=6, value=unit_price)
        ws.cell(row=row_num, column=7, value=total_amount)
        ws.cell(row=row_num, column=8, value=order_date.strftime("%Y-%m-%d"))
        ws.cell(row=row_num, column=9, value=region)
        ws.cell(row=row_num, column=10, value=status)
        ws.cell(row=row_num, column=11, value=payment_method)
        ws.cell(row=row_num, column=12, value=shipping_cost)
    
    # Adjust column widths
    column_widths = {
        'A': 12, 'B': 18, 'C': 16, 'D': 18, 'E': 10, 'F': 12,
        'G': 14, 'H': 12, 'I': 10, 'J': 12, 'K': 18, 'L': 14
    }
    
    for column, width in column_widths.items():
        ws.column_dimensions[column].width = width
    
    # Save the workbook
    filename = "test_data_sales.xlsx"
    wb.save(filename)
    print(f"✓ Test Excel file created: {filename}")
    print(f"  - {len(headers)} columns")
    print(f"  - 50 rows of sample data")
    print(f"  - File location: {filename}")
    
    return filename

if __name__ == "__main__":
    try:
        create_test_excel()
    except ImportError:
        print("❌ Error: openpyxl library not found")
        print("Install it with: pip install openpyxl")
    except Exception as e:
        print(f"❌ Error creating Excel file: {e}")
